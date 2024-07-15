import os
from time import perf_counter
from glob import glob
from itertools import product
from typing import List, Union

import torch
import numpy as np
import pandas as pd
from xarray import open_dataset
from h5py import File as File_hdf
from openpyxl import load_workbook, Workbook
from pandas import DataFrame, MultiIndex

from .constant import (
    ACTIVE_PERF_COUNTER, 
    DATA_DIRECTORY
)

from .preprocessing import min_max_scale, generate_synthetic_signal, generate_noisy_signals

def report_time(func):
    """
    Wrapper that prints func execution time 
    *measured with time.perf_counter
    """
    def wrapper(*args, **kwargs):
        if ACTIVE_PERF_COUNTER:
            start = perf_counter()
        results = func(*args, **kwargs)
        if ACTIVE_PERF_COUNTER:
            t = round(perf_counter() - start, 3)
            print(f"({func.__name__}) Execution time: {t}")
        return results
    return wrapper

def report_end(func):
    """
    Wrapper that prints termination of function
    """
    def wrapper(*args, **kwargs):
        results = func(*args, **kwargs)
        print(f"({func.__name__}) done with param 1: {str(args[0])}")
        return results
    return wrapper

def try_chdir(path_name: str=""):
    """
    Safely executes a change directory command
    """
    if not path_name:
        path_name = DATA_DIRECTORY
    try:
        os.chdir(path_name)
    except:
        print(f"Directory '{path_name}' not found")

###################
#### MODEL I/O ####
###################

def load_model(model_placeholder, path, on_eval=False):
    model_placeholder.load_state_dict(torch.load(path))
    if on_eval:
        model_placeholder.eval()
    return model_placeholder

def load_model_state_dict(path, on_eval=False):
    model = torch.load(path)
    if on_eval:
        model.eval()
    return model

##################
#### Data I/O ####
##################

@report_time
def read_impedances(data_path="", on_array: bool=True, **kwargs) -> np.ndarray:
    """
    Reads all the .xlsx files from the swept tests directory 
    and arranges the data into shape (load, sweep, sensor, row, col)
    
    Returns:
        A NumPy array containing the data extracted from the .xlsx files.
    """
    file_regex = kwargs.get("file_regex", "*.xlsx")
    version_sep = kwargs.get("version_sep", "_")
    file_format = kwargs.get("file_format", "xlsx")

    dataset = []
    path = os.path.join(data_path, file_regex)
    for f_name in sorted(glob(path), key=lambda x: int(x.split(".")[0].split(version_sep)[-1])):
        print(f"Reading: {f_name}")
        if not f_name.endswith(file_format):
            continue
        dataset.append(_read_sheet(f_name, **kwargs))
    if on_array:
        return np.array(dataset)
    return dataset

def write_hdf5_file(
    f_name="", 
    dataset_name="", 
    data=None, 
    strict_write=True
) -> bool:
    """
    Stores an input NdArray into an HDF5 dataset and writes a file.
    """
    if not f_name or not dataset_name or data is None:
        raise Exception("Input error. 'f_name', 'dataset_name', or 'data' is not defined.")
    if os.path.exists(f_name):
        print(f"The file {f_name} already exists")
        return False
    mode = "w-" if strict_write else "w"
    with File_hdf(f_name, mode) as f:
        f.create_dataset(dataset_name, data=data)
    return True

def read_hdf5_file(
    f_name="", 
    dataset_name=""
) -> np.ndarray:
    """
    Reads an input dataset from a HDF5 file.
    """
    if not f_name or not dataset_name:
        raise Exception("Input error. 'f_name' or 'dataset_name' is not defined.")
    with File_hdf(f_name, 'r') as f:
        return np.array(f[dataset_name])

# ''Private'' functions
@report_time
def _extract_sheet_data(
    ws: Workbook,
    **kwargs
) -> List[List[float]]:
    """
    Double-loops through a spreadsheet sheet and extracts the values.
    Inputs:
        - ws: A Workbook instance representing the spreadsheet sheet.
    Returns:
        A double-nested list containing the values extracted from the sheet.
    """
    if not ws:
        raise Exception("Input error. 'ws' is not defined.")
    data = []
    matrix_rows = []
    for row in ws.iter_rows(values_only=True):
        if row and isinstance(row[0], (int, float)):
            matrix_rows.append(row)
            continue
        if matrix_rows:
            data.append(_stack_features_and_ranges(matrix_rows, **kwargs))
            matrix_rows = []
    if matrix_rows:
        data.append(_stack_features_and_ranges(matrix_rows, **kwargs))
    return np.array(data)

@report_end
def _read_sheet(
    f_name: str,
    **kwargs
) -> Union[List[List[List[float]]], np.ndarray]:
    """
    Reads a valid sheet from a spreadsheet which contains data from sweeps.
    
    A valid sheet name is defined as 'Sweep_i', where 'i' is an integer.
    
    Inputs:
        - f_name: The name of the .xlsx file.
        - **kwargs: Additional keyword arguments.
        
    Returns:
        If on_array is True, returns a NumPy array of sheet names. 
        If on_array is False, returns a list of sheet names.
    """
    # Load the Excel workbook
    work_book = load_workbook(f_name, 
                              read_only=True,
                              data_only=True
                             )
    # Access the active worksheet
    sheets_data = []
    for sheet_name in work_book.sheetnames:
        if sheet_name[-1].isdigit():
            sheets_data.append(
                _extract_sheet_data(work_book[sheet_name], **kwargs)
            )
    return np.array(sheets_data)

def _stack_features_and_ranges(data, **kwargs):
    """
    Stacks features and ranges of data.
    
    Inputs:
        - data: Input data.
        - **kwargs: Additional keyword arguments.
        
    Returns:
        Stacked features and ranges of the data.
    """
    n_features = kwargs.get("n_features", 3)
    stack_ranges = kwargs.get("stack_ranges", True)
    data_array = np.array(data)
    stack = []
    for i in range(0, len(data_array[0]), n_features):
        stack.append(data_array[:, i:i+n_features])
    if stack_ranges:
        return np.vstack(stack)
    return np.array(stack)

def read_and_format_to_dataframe(data_path, stack_ranges=True, columns_names=None, index_names=None):
    """
    Reads data from a specified path and formats it into a DataFrame.
    
    Inputs:
        - data_path: Path to the data.
        - stack_ranges: Boolean indicating whether to stack ranges.
        - columns_names: Names of the columns.
        - index_names: Names of the indices.
        
    Returns:
        Formatted DataFrame.
    """

    if columns_names is None:
        column_names = ["frequency", "real", "imaginary"]
    
    # Read the data, parsed it into an nd-array
    data = read_impedances(data_path=data_path, 
                           stack_ranges=stack_ranges)

    # Set default index names
    if index_names is None:
        if stack_ranges:
            index_names = ["load", "sweep", "sensor", "freq_step"]
        else:
            index_names = ["load", "sweep", "sensor", "freq_range", "freq_step"]
    
    if stack_ranges:
        n_loads, n_sweeps, n_sensors, n_steps, n_vars = data.shape
        
        # Combination of 4 indices
        combs = product(range(n_loads),
                        range(n_sweeps),
                        range(n_sensors),
                        range(n_steps)
                       )
    else:
        n_loads, n_sweeps, n_sensors, n_ranges, n_steps, n_vars = data.shape
        
        # Combination of 5 indices
        combs = product(range(n_loads),
                        range(n_sweeps),
                        range(n_sensors),
                        range(n_ranges),
                        range(n_steps)
                       )
    
    # Make dataframe (reshape data into a 2d-array and make multi_index)
    return DataFrame(data.reshape(-1, n_vars), 
                     index=MultiIndex.from_tuples(list(combs), 
                                                  names=index_names), 
                     columns=column_names
                    )

def read_xarray_dataset(path, engine="netcdf4", drop_dups=False):
    """
    Reads a dataset from a specified path using xarray.
    
    Inputs:
        - path: Path to the dataset.
        - engine: Engine for opening the dataset.
        
    Returns:
        Data, columns, number of steps, and data variables.
    """
    data = open_dataset(path, engine=engine)
    if drop_dups:
        _, unique_steps = np.unique(data["frequency"], return_index=True)
        data = data.sel(step=unique_steps)
    
    columns = list(data.sizes.keys())
    #ns = list(data.sizes.values())
    data_vars = list(data.data_vars.keys())
    
    print(f"Shape of dimensions:\n\t{data.dims}")
    print(f"Shape of variables:\n\t{data.data_vars}")

    #return data, columns, ns, data_vars
    return data, columns, data_vars

def build_array(x, values, dim, main_feature="real", permutations=None, n_splits=0, add_noise_augmentation=False, add_minmax_augmentation=False, probabilities_to_positive=None, axis_agg=-1, bound_to_positives=True, on_load_target=False, on_ids=False, final_reshape=None, on_squeeze_target=True):
    """
    Builds an array of signals for processing.

    Parameters:
        - x: Dataset containing the impedance data (samples, sweeps, categories (sensors), steps).
        - values: Values of the specified dimension.
        - dim: Dimension along which to select values.
        - main_feature: Main feature to use.
        - permutations: Permutations to apply to the data.
        - n_splits: Number of splits to apply to the sequence.
        - add_noise_augmentation: Whether to add noise augmentation.
        - add_minmax_augmentation: Whether to add min-max augmentation.
        - probabilities_to_positive: Whether to add probabilities.
        - axis_agg: Axis along which to aggregate the data.
        - bound_to_positives: Whether to clip negative values to positive.
        - on_load_target: Whether to return a target of integer labels.
        - on_load_target: Whether to return an array of IDs for each sample.
        - final_reshape: tuple with reshape dims to return. 

    Returns:
        Processed data signals and frequency encoding.
    """
    # Clip real impedance to positive values
    data = x[main_feature].clip(0.0, None).sel({dim: values}).values
    n_loads, n_sweeps, n_categories, n_steps = data.shape

    # Initialize list of augmentations
    augmentations = []
    
    # Apply min-max augmentation before normalization
    if add_minmax_augmentation:
        print("\nRunning min-max augmentation:")
        print("\tShape of input data", data.shape)
        
        # Before (load, sweeps, categories, steps)
        x_minmax_aug = generate_synthetic_signal(data, probabilities_to_positive=probabilities_to_positive)
        # After (n_probabilities=total_examples, load, categories, steps)
        augmentations.append(x_minmax_aug)
        
        print("\tShape of output data:", x_minmax_aug.shape)
    
    # Apply white noise augmentation after normalization
    if add_noise_augmentation:
        print("\nRunning noise augmentation:")
        print("\tShape of input data:", data.shape)
        
        x_noise_aug = generate_noisy_signals(data)
        # Transpose dimension load and sweeps stack dims 0 and 1 (sweeps): 
        # Before (examples, load, sweeps, categories, steps) 
        x_noise_aug = x_noise_aug.transpose(0, 2, 1, 3, 4)
        
        examples_noise, examples_sweep, *shape_rest = x_noise_aug.shape
        x_noise_aug = x_noise_aug.reshape(-1, *shape_rest)
        # After 1 (examples, sweeps, load, , categories, steps)
        # After 2 (examples * sweeps=total_examples, load, categories, steps)
        print("\tShape of output data:", x_noise_aug.shape)
        augmentations.append(x_noise_aug)

    # Transpose dims load and sweep.
    # Before (load, sweeps, categories, steps)
    data = data.transpose(1, 0, 2, 3)
    # After (sweeps, load, categories, steps)
    
    print("\nShape of the original data (after 0 (load) and 1 (sweeps) dims transposition)):", data.shape)

    # Stack first dimension of original and augmentations data
    augmented_data = np.vstack([data, *augmentations])

    if n_splits > 0:
        # Modify steps that the sequence is splitted in 9 ranges mixed within signal axis (axis=0)
        n_steps =  n_steps // n_splits
        *shape_rest, _ = augmented_data.shape
        # Split the steps dim and spread the data along the new dim 'splits' and 'step' dim
        augmented_data = augmented_data.reshape(*shape_rest, n_splits, n_steps)
        # Transpose the dims so that 'split' dim is at 0 and the left ones are rolled 1 position to the right:
        # Example: (samples 0, load 1, sensor 2, splits 3, step 4) ->  (splits 3, samples 0, load 1 , sensor 2, step 4)
        augmented_data = augmented_data.transpose(3, 0, 1, 2, 4)
    
    if permutations:
        augmented_data = np.transpose(augmented_data, axes=permutations)
        
    # Apply min-max normalization
    augmented_data = min_max_scale(augmented_data, axis=-1)
    
    # cast to tensor
    augmented_data = torch.tensor(augmented_data, dtype=torch.float32)
    
    if on_load_target:
        target = make_target_load(augmented_data, on_squeeze_target=on_squeeze_target)
        
    if on_ids:
        ids = make_signal_ids(augmented_data)

    # Reshape to 2D tensor (samples/signals, step)
    if final_reshape is not None:
        augmented_data = augmented_data.reshape(final_reshape)
    
    print("\nShape of the augmented data:", augmented_data.shape)

    # Return options
    if on_load_target and on_ids:
        return augmented_data, target, ids
    elif on_load_target:
        return augmented_data, target
    elif on_ids:
        return augmented_data, ids
    else:
        return augmented_data
    
def make_target_load(data, on_squeeze_target=True):
    """
    Computes a target of integer labels from data Ndarray.

    Parameters:
        - data (ndarray, (*sample, load, sensor, step)): Number of loads in the dataset.
        
    Returns:
        1D torch tensor.
    """
    
    # Skip dim 'load' to compute the number of examples
    *rest, n_loads, n_sensors, n_steps = data.shape
    n_examples = np.prod(rest) * n_sensors
    
    target_load = torch.tensor(
        np.repeat([x for x in range(n_loads)], repeats=n_examples), 
        dtype=torch.int32
    )

    if on_squeeze_target:
        return target_load.squeeze().to(torch.int64)
    
    # Reshape from (n) to (n, 1) and assign proper dtype
    return target_load.unsqueeze(-1).to(torch.int64)

def make_one_hot_encoding(labels, n_categories=None):
    """
    Make one-hot encoding from a vector of integer labels.
    """
    if n_categories is None:
        n_categories = len(np.unique(labels))
    
    # Expanded matrix of categories in sequential order (0, 1, 2, ... n- 1)
    categories_expanded_mat = np.repeat(np.arange(n_categories).reshape(1, -1), n_examples, axis=0)

    oh_matrix = np.equal(categories_expanded_mat, labels).astype(int)
    return oh_matrix

def test_make_one_hot(oh, labels):
    positions_dummy = oh_matrix.argmax(axis=1)
    return (positions_dummy == labels.squeeze()).all()

def make_signal_ids(data, names=None):
    """
    Builds a Tensor of categories which serves as id for each signal.

    Parameters:
        - data (ndarray, (*sample, load, sensor, step))
          or
          (ndarray, (split, *sample, load, sensor, step)): original data array.

    Returns:
        2D Tensor (ndarray) of categories. Each row.
    """
    from itertools import product
    import pandas as pd
    
    if names is None:
        names = ["split", "load", "sensor"]

    *shape_rest, n_loads, n_sensors, n_steps = data.shape
    load_vector = np.arange(n_loads)
    sensor_vector = np.arange(n_sensors)
    
    if len(shape_rest) > 1:
        n_splits, n_examples = shape_rest
        split_vector = np.arange(n_splits)
        
        base_combinations = product(split_vector, load_vector, sensor_vector)
    else:
        n_examples = np.prod(shape_rest)

        base_combinations = product(load_vector, sensor_vector)

    combinations = n_examples * list(base_combinations)
    
    return pd.MultiIndex.from_tuples(combinations, names=names)